import numpy as np
import openpyxl
import dateutil
from ipdb import set_trace as stop
import datetime
import pickle
import pystan
from hashlib import md5

def stan_cache(model_code, model_name=None, **kwargs):
    """Use just as you would `stan`"""
    code_hash = md5(model_code.encode('ascii')).hexdigest()
    if model_name is None:
        cache_fn = 'cached-model-{}.pkl'.format(code_hash)
    else:
        cache_fn = 'cached-{}-{}.pkl'.format(model_name, code_hash)
    try:
        sm = pickle.load(open(cache_fn, 'rb'))
    except:
        sm = pystan.StanModel(model_code=model_code)
        with open(cache_fn, 'wb') as f:
            pickle.dump(sm, f)
    else:
        print("Using cached StanModel")
    return sm.sampling(**kwargs)

def toenglish(s):
    spanish = ['ene', 'abr', 'ago', 'dic', 'de mayo de']
    english = ['jan', 'apr', 'aug', 'dec', 'may']
    for (j, month) in enumerate(spanish):
        s = s.replace(month, english[j])
    return s

def getPercentage(s):
    if (s[0] not in ['0','1','2','3','4','5','6','7','8','9']):
        return 0
    else:
        if (s.find('%') != -1):
            return float(s.split('%')[0].replace(',','.')) / 100.0
        else:
            return float(s.split('\n')[0].replace(',','.')) / 100.0

def getSigma(s):
    left = s.find('(')
    right = s.find(')')    
    if (s[left+1:right] in ['?', '-']):
        return 0.03
    else:
        return 1.0 / np.sqrt(float(s[left+1:right]))

def weeksDifference(d1, d2):
    monday1 = (d1 - datetime.timedelta(days=d1.weekday()))
    monday2 = (d2 - datetime.timedelta(days=d2.weekday()))

    return int((monday2 - monday1).days / 7)

wb = openpyxl.load_workbook("data/sondeos2015.xlsx")
ws = wb.active

empresas = ['GAD3', 'Encuestamos', 'GESOP', 'Metroscopia', 'Celeste-Tel',' Demoscopia Servicios', 'Simple Lógica', 'CIS', 'TNS Demoscopia', 'Invymark', 'Resultados de las elecciones']

empresaSondeoAll = []
sondeosAll = []
dateAll = []
sigmaAll = []

otrosPartidos = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'R']

nSondeos = 302
for i in range(100):#nSondeos):
    empresa = ws['A{0}'.format(i+2)].value
    for (loop, emp) in enumerate(empresas):
        if (empresa.find(emp) != -1):
            empresaSondeo = loop

    if (empresaSondeo == 10):
        sigma = 0.0001
    else:
        sigma = getSigma(empresa)

    PP = getPercentage(ws['C{0}'.format(i+2)].value)
    PSOE = getPercentage(ws['D{0}'.format(i+2)].value)
    IU = getPercentage(ws['E{0}'.format(i+2)].value)
    PODEMOS = getPercentage(ws['P{0}'.format(i+2)].value)
    CS = getPercentage(ws['Q{0}'.format(i+2)].value)

    total = PP + PSOE + IU + PODEMOS + CS
    otros = 1.0 - total    

    tmp = ws['B{0}'.format(i+2)].value
    if (isinstance(tmp, datetime.date)):
        date = tmp
    else:
        date = dateutil.parser.parse(toenglish(ws['B{0}'.format(i+2)].value.split('-')[-1].lower()))    

    tmp = date.year + (date.month-1.0) / 12.0

    sondeo = [PP, PSOE, IU, PODEMOS, CS, otros]

    sondeosAll.append(sondeo)
    sigmaAll.append(sigma)
    
    dateAll.append(date)
    empresaSondeoAll.append(empresaSondeo+1)

    print ("{0} - {1} {8} - sigma={9} : PP={2} - PSOE={3} - IU={4} - PODEMOS={5} - CS={6} - Resto={7}".format(i, 
        empresas[empresaSondeo], PP*100, PSOE*100, IU*100, PODEMOS*100, CS*100, otros*100, date, sigma*100))

sondeosAll = np.array(sondeosAll)

nSondeos, nPartidos = sondeosAll.shape
nEmpresas = len(empresas)

# Compute week of every poll
weekAll = []
for i in range(nSondeos):
    weekAll.append(weeksDifference(dateAll[nSondeos-1], dateAll[i]) + 1)

nDates = max(weekAll)

# Reverse all lists
sondeosAll = sondeosAll[::-1]
empresaSondeoAll.reverse()
weekAll.reverse()
sigmaAll.reverse()

dictionary = {'NPartidos': nPartidos, 'NSondeos': nSondeos, 'NEmpresas': nEmpresas, 'NDates': nDates, 'empresa': empresaSondeoAll, 'sondeos': sondeosAll, 
'date': weekAll, 'sigmaSondeo': sigmaAll, 'alpha': np.ones(nPartidos)*0.3}

f = open('model.stan', 'r')
model = f.read()
f.close()

out = stan_cache(model, model_name='elecciones', data=dictionary, chains=1)

